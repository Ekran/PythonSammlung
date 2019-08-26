# created by eschreiter@hszg.de
# released under GPL-3.0 https://www.gnu.org/licenses/gpl-3.0.de.html
# 2017-05-29 started
# 2018-03-13 alpha --> exports ics for students, Staff, but some dates are missing
# 2019-08-26 Vollständigkeit des Exports der ICS Daten ist ungeprüft

# Idee: Eine Excel-Tabelle einlesen die auf mehreren Seiten Praktikumstermine enthält
# primäres Ziel:
# - ical Datei exportieren
# - einmal pro Matrikelgruppen
# - einmal pro Betreuername
# - ics Datei kann in Kalenderprogramm eingelesen werden, z.B. Thunderbild, Groupwise
# - empfohlene Arbeitsweise: ics Import in Extral Kalender "Praktikum" o.ä. einlesen
# - vor oder bei erneutenm Import alten Inhalt löschen, zumindest den in dem Semester (Monatsansicht, alles markieren, löschen) oder Kalender löschen
# offen für Zukunft
# - in Verzeichnisse einlesen und diff auf ical durchführen = hat sich etwas geändert --> Benachrichtigung (per email)
# - Statistik, welches Matrikel, Raum, Personal hat am meisten

## Quellen:
# https://openpyxl.readthedocs.io/en/default/usage.html#read-an-existing-workbook

## installieren
# pip install openpyxl

## import
from openpyxl import load_workbook
import datetime
import csv # für export_termintabelle
import icalendar # für export_ical
# from datetime import datetime # für export_ical
import pytz # für export_ical
import tempfile, os # für export_ical


def filter_termintabelle(Termintabelle,typ,text):
# Filtert Terminintabelle nach typ, text enthält Namen
## 1 - Matrikel
## 2 - Laboring
## 3 - Raum
## 4 - Prof
## 5 - Gruppe
# Termintabelle.append([ Matrikel, Gruppe, Versuchsname, Datum_Uhrzeit, Ort, Laboring, Prof])
    Termintabelle_gefiltert = list()
    for termin in Termintabelle:
        found = 0
        if typ == 1 and str(termin[0]).find(str(text))==0: # Matrikel
            found = 1
        # Laboring, Prof
        if (typ == 2 or typ == 4) and (str(termin[5]).find(str(text))>=0 or str(termin[6]).find(str(text))>=0):
            found = 1
        # Raum
        if typ == 3 and str(termin[4]).find(str(text))>=0:
            found = 1
        # Gruppe
        if typ == 5 and int(termin[1]) == int(text):
            found = 1
    
        if found != 0:
           Termintabelle_gefiltert.append(termin)
    return Termintabelle_gefiltert

def get_profs(Termintabelle):
    Profs = list()
    # Termintabelle.append([ Matrikel, Gruppe, Versuchsname, Datum_Uhrzeit, Ort, Laboring, Prof])
    for termin in Termintabelle:
        if termin[6] not in Profs:
            Profs.append(termin[6])
    return Profs

def get_laborings(Termintabelle):
    Laborings = list()
    for termin in Termintabelle:
        if termin[5] not in Laborings:
            Laborings.append(termin[5])    
    return Laborings

def get_Raeume(Termintabelle):
    Raeume = list()
    for termin in Termintabelle:
        if termin[4] not in Raeume:
            Raeume.append(termin[4])    
    return Raeume


def export_ical(Termintabelle,file_out):
   #http://icalendar.readthedocs.io/en/latest/usage.html
    cal = icalendar.Calendar()
    cal.add('prodid', '-//My calendar product//mxm.dk//')
    cal.add('version', '2.0')
    for termin in Termintabelle:
        # Termintabelle.append([ Matrikel, Gruppe, Versuchsname, Datum_Uhrzeit, Ort, Laboring, Prof])
        event = icalendar.Event()
        Titel = str(termin[0]) + " Gr." + str(termin[1]) + " " + str(termin[2])+ " DI " + str(termin[5])+ "/Prof. " + str(termin[6])
        event.add('summary', Titel)
        event.add('dtstart', termin[3])
        event.add('dtend', termin[3] +  datetime.timedelta(minutes=120)) # https://stackoverflow.com/questions/100210/what-is-the-standard-way-to-add-n-seconds-to-datetime-time-in-python
        #event.add('dtend', datetime(2018,3,12,10,0,0,tzinfo=pytz.utc))
        event.add('dtstamp', datetime.datetime.now())
        event.add('location', termin[4])
        cal.add_component(event)
    f = open(file_out, 'wb')
    f.write(cal.to_ical())
    f.close()
    

def export_termintabelle(Termintabelle,file_out_csv):
## https://stackoverflow.com/questions/2084069/create-a-csv-file-with-values-from-a-python-list
    with open(file_out_csv, 'w') as myfile:
        wr = csv.writer(myfile, quoting=csv.QUOTE_ALL)
        for termine in Termintabelle:
            wr.writerow(termine)





def combine_Versuche_Termine(Versuche, Termine):
    # Termintabelle hat folgende Felder
     # Matrikel
     # Gruppe
     # Versuchsname
     # Datum_Uhrzeit
     # Ort
     # Laboring
     # Prof
    # Dazu wird über die Zeilen der 'Termine' iteriert und
    # jeweils aus 'Versuche' die Infos geholt
    Termintabelle = list()
    for Termin in Termine:
        # print(str(Termin) + str(Termin[0]))
        Matrikel = Termin[0]
        Gruppe = Termin[1]
        Datum_Uhrzeit = Termin[2]
        Versuchsname = Termin[3]
        for Versuch in Versuche:
            if (Versuch[0] == Matrikel) and (str(Versuch[1]) == str(Versuchsname)):
                print(Termin)
                print(Versuch)
                Versuchsname += ' ' + Versuch[2]
                Ort = Versuch[3]
                Namen = Versuch[4]
                Prof, Laboring = Namen.split('/',1)
                Laboring = Laboring.replace("DI","").replace("Dr","").replace("DM","").strip(" ,.")
                Prof = Prof.replace("Prof","").strip(" ,.")
                Termintabelle.append([ Matrikel, Gruppe, Versuchsname, Datum_Uhrzeit, Ort, Laboring, Prof])
                break # for Versuche ...
        
    return Termintabelle

def test_Zeile_ist_Termin(ws,zeile,spalte):
    inhalt = ws[Zell_Addresse(zeile,spalte)].value
    if not isinstance(inhalt, str):
        return 0
    if inhalt.find('.')>=0 and inhalt.find(',')>=0: # zelle enthält Zeitstempel im Format '25.01., 16.15'
        return 1
    return 0

def get_Datum_Zeit(Zeit_als_Text):
    if not isinstance(Zeit_als_Text, str):
        return 0
    Zeit_als_Text = Zeit_als_Text.replace(' ', '')
    Zeit_als_Text = Zeit_als_Text.replace(',', '')
    print(Zeit_als_Text)
    Monat = 0
    Tag = 0
    Stunde = 0
    Minute = 0
    Tag, Monat, Stunde, Minute = Zeit_als_Text.split('.')
    return int(Monat), int(Tag), int(Stunde), int(Minute)


def get_Anzahl_Termine(ws,zeile_anfang_termine):
    Anzahl_Termine = -1
    Anzahl_maximale_Zeilen = 30
    i = 0
    if test_Zeile_ist_Termin(ws,zeile_anfang_termine+i,1):
        Anzahl_Termine = 1
        i = i + 1
    while test_Zeile_ist_Termin(ws,zeile_anfang_termine+i,1) or test_Zeile_ist_Termin(ws,zeile_anfang_termine+i+1,1):
        i = i + 1
        Anzahl_Termine = Anzahl_Termine + 1
        if i > Anzahl_maximale_Zeilen:
            break
    return Anzahl_Termine

def get_einzelDatum_aus_Bereich(inhalt):
    # aus Text in der Art '13.-16.11.' oder '13.11.-16.11.'
    # mehrere Termine erzeugen
    Monat = list()
    Tag = list()
    Stunde = list()
    Minute = list()
    Mehrere_Daten = list()
    Tag1 = 0
    Tag2 = 0
    Monat1 = 0
    Monat2 = 0
    inhalt = inhalt.replace(' ','') # Leerzeichen entfernen
    inhalt = inhalt.replace(',','') # Kommas entfernen
    if inhalt.count('-') is not 1:
        print('Fehler in get_einzelDatum_aus_Bereich' + str(inhalt))
        return list() ## leere Liste
    inhalt1, inhalt2 = inhalt.split('-')
    split1 = inhalt1.split('.')
    split2 = inhalt2.split('.')
    Tag1 = int(split1[0])
    if len(split1[1]):
        Monat1 = int(split1[1])
    Tag2 = int(split2[0])
    if len(split2[1]):
        Monat2 = int(split2[1])
    for n in range(Tag1, Tag2):
        Mehrere_Daten.append([Monat2,n,0,0])
    return Mehrere_Daten

def get_Termine(wb, sheet_namen,Gruppen,zeile_anfang_termine):
    Termine = list()

    Datum_Uhrzeit = datetime.datetime.now()
    Datum_Uhrzeit = Datum_Uhrzeit.replace(second = 0, microsecond = 0)
    Monat, Tag, Stunde, Minute = 0,0,0,0
    Gruppe = 0
    Versuch = ''

    j=0
    for name in sheet_namen:
        ## erste Tabelle auswählen
        ws = wb[name]
        # Tabelle suchen
        #[Zeile_Start, Spalte_Versuchsnummer, Spalte_Versuchsname, Spalte_Ort, Spalte_Verantwortlicher, Anzahl_Zeilen] = get_Position_Versuchsliste(ws)
        Anzahl_Zeilen = get_Anzahl_Termine(ws,zeile_anfang_termine[j])
        print('Suche nach Terminen in Tabellenblatt ' + name + ' mit ' + str(Gruppen[j]) + ' Gruppen und ' + str(Anzahl_Zeilen) + ' Zeilen Terminen')
        # Werte einlesen wenn Zeilen gefunden wurden
        if Anzahl_Zeilen < 1:
            print('Keine Termine in ' + name + ' ab Zeile ' +str(zeile_anfang_termine[j]) + '?')
            continue # nächste Tabellenseite
        for i in range(0,Anzahl_Zeilen): # über alle Zeilen mit Terminen
            # erste Spalte enthält Datum
            inhalt = ws[Zell_Addresse(zeile_anfang_termine[j]+i,1)].value
            if not isinstance(inhalt, str): # Wenn leer:
                print('Terminfindung Datum leer!? ' + name + ' Zeile ' +str(zeile_anfang_termine[j]+i))
                continue
            # ToDo: suche nach '-' --> Datumsbereich z.B. '24.10.-27.10.' --> in Einzelganztagstermine verwandeln
            if inhalt.find('-') >= 0:
                Mehrere_Daten = get_einzelDatum_aus_Bereich(inhalt)
                for einzeldatum in Mehrere_Daten:
                    inhalt = ws[Zell_Addresse(zeile_anfang_termine[j]+i,2)].value # erste Spalte Versuchsname lesen
                    for n in range(1,Gruppen[j]+1): #für jede Gruppe eintragen
                        Versuchsname = str(inhalt)
                        Gruppe = int(n)
                        Datum_Uhrzeit = Datum_Uhrzeit.replace(month=einzeldatum[0], day=einzeldatum[1], hour=0, minute=0)
                        print([name,Gruppe,Datum_Uhrzeit, Versuchsname])
                        Termine.append([name,Gruppe+1,Datum_Uhrzeit, Versuchsname])
                continue # nächste Zeile

            if not inhalt.count('.') == 3:
                print('Datum hat falsches Format')
                continue
            Monat, Tag, Stunde, Minute = get_Datum_Zeit(inhalt)
            if int(Monat) or int(Tag): # beides kann nicht null sein
                Datum_Uhrzeit = Datum_Uhrzeit.replace(month=Monat, day=Tag, hour=Stunde, minute=Minute)
                print('Datum gefunden')
            else:
                print('Datum kaputt?')
                continue
            for n in range(1,Gruppen[j]+1):
                # Versuchsnummer#       Versuchsname  #      Ort   #     Verantwortlicher
                # pro Zeile über eine Spalte pro Zeile prüfen ob nicht leer
                inhalt = ws[Zell_Addresse(zeile_anfang_termine[j]+i,n+1)].value
                if (inhalt is  None): # Wenn leer:
                    print('kein Versuch für Gruppe ' + str(int(n)) + ' eingetragen in ' + str(Zell_Addresse(zeile_anfang_termine[j]+i,n+1)))
                    continue
                # print(str(inhalt))
                Versuchsname = str(inhalt)
                Gruppe = int(n)
                print([name,Gruppe,Datum_Uhrzeit, Versuchsname])
                Termine.append([name,Gruppe,Datum_Uhrzeit, Versuchsname])
        j = j + 1
    return Termine


def Zell_Addresse(Zeile,Spalte):
    # Wandelt Nummern für Zeile und Spalte in Excel-Zell-Addressen um, z.B. 'A1' oder 'AD123'
    Zeile = int(Zeile)
    Spalte = int(Spalte)
    Addresse = ''
    if Spalte > 26:
        Addresse = chr(ord('A')+Spalte//26-1)
        Spalte = Spalte % 26 + 1
    Addresse = Addresse + chr(ord('A')+Spalte-1)
    Addresse = Addresse + str(Zeile)
    return Addresse

def get_Position_Versuchsliste(ws):
    # suche nach 'Nr.' oder 'Bezeichnung'
    # dann ob und wo daneben 'Versuch', V.raum, 'Lehrfachverantw.' steht
    zeile_suchen = 1
    zeile_max = 50
    spalte_suchen = 1
    spalte_max = 24
    gefunden = 0
    Zeile_Start = 0
    Spalte_Versuchsnummer = 0
    Spalte_Versuchsname = 0
    Spalte_Ort = 0
    Spalte_Verantwortlicher = 0
    while True:
        if gefunden:
            break
        if spalte_suchen > spalte_max:
            zeile_suchen = zeile_suchen + 1
            spalte_suchen = 1
            continue
        if zeile_suchen > zeile_max:
            break # am Ende angekommen!
        #print('Suchen in ' + Zell_Addresse(zeile_suchen,spalte_suchen))
        inhalt = ws[Zell_Addresse(zeile_suchen,spalte_suchen)].value
        ## jetzt Zelle verarbeiten
        if isinstance(inhalt, str):
            # Zelle enthält Text!
            if inhalt.find('Nr.') >= 0 or inhalt.find('Bezeichnung') >= 0:
                print('Zeile: ' + str(zeile_suchen))
                # in Zeile nach hinten suchen nach 'Versuch'
                Zeile_Start = zeile_suchen + 1
                Spalte_Versuchsnummer = spalte_suchen
                print('Spalte_Versuchsnummer ' + str(Spalte_Versuchsnummer))
                spalte = spalte_suchen + 1
                while True:
                    if spalte_suchen > spalte_max:
                        break
                    #print('in Zelle suchen ' + Zell_Addresse(zeile_suchen,spalte_suchen))
                    inhalt = ws[Zell_Addresse(zeile_suchen,spalte_suchen)].value
                    if isinstance(inhalt, str):
                        if inhalt.find('Versuch')>=0 and not inhalt.find('Versuchsraum')>=0:
                            Spalte_Versuchsname = spalte_suchen
                            print('Spalte_Versuchsname ' + str(Spalte_Versuchsname))
                        if (inhalt.find('V.Raum')>=0 or inhalt.find('V.raum')>=0 or inhalt.find('Versuchsr')>=0):
                            Spalte_Ort = spalte_suchen
                            print('Spalte_Ort' + str(Spalte_Ort))
                        if inhalt.find('Lehrfachver')>=0:
                            Spalte_Verantwortlicher = spalte_suchen
                            print('Spalte_Verantwortlicher ' + str(Spalte_Verantwortlicher))
                            if Spalte_Verantwortlicher > 0:
                                print('Variable ist gesetzt')
                            if ((Spalte_Verantwortlicher > 0) and (Spalte_Ort > 0)):
                                print('Variablen sind gesetzt')
                        if (Spalte_Versuchsname > 0 and (Spalte_Ort  > 0) and  Spalte_Verantwortlicher  > 0):
                            # alle Angaben gefunden!
                            print('Alles gefunden')
                            gefunden = 1
                            break # while über Spalten
                        if spalte_suchen >= spalte_max:
                            # nichts/nicht alles gefunden!
                            print('Spaltenmax erreicht, nicht alles gefunden!')
                            break
                        # weitersuchen
                        spalte_suchen = spalte_suchen + 1
                        continue
##### Abbruchbedingung für maximale Zeile und Spalte!
                    else: # kein Text
                        spalte_suchen = spalte_suchen + 1
                        continue
                    spalte_suchen = spalte_suchen + 1
            else:
                # String enthält nicht gesuchten Text
                # --> weitersuchen
                spalte_suchen = spalte_suchen + 1
                continue
        else: # kein String!
            # --> weitersuchen
            spalte_suchen = spalte_suchen + 1
            continue

    # Anzahl der Zeilen suchen
    # erste leere Zeile nach unten suchen!
    #, , , , , Anzahl_Zeilen
    zeile = zeile_suchen
    while True:
        inhalt = ws[Zell_Addresse(zeile,spalte_suchen)].value
        if isinstance(inhalt, str):
            zeile = zeile + 1
            continue
        break
    Anzahl_Zeilen = zeile - zeile_suchen - 1
    print('Anzahl Zeilen ' + str(Anzahl_Zeilen))

    print([Zeile_Start, Spalte_Versuchsnummer, Spalte_Versuchsname, Spalte_Ort, Spalte_Verantwortlicher, Anzahl_Zeilen])
    return [Zeile_Start, Spalte_Versuchsnummer, Spalte_Versuchsname, Spalte_Ort, Spalte_Verantwortlicher, Anzahl_Zeilen]


def get_Versuche(wb, sheet_namen):
    Versuche = list()

    Versuchsnummer = 0
    Versuchsname = ''
    Ort = ''
    Verantwortlicher = ''

    for name in sheet_namen:
        ## erste Tabelle auswählen
        ws = wb[name]
        print('Suche nach Versuchen in Tabellenblatt ' + name)
        # Tabelle suchen
        [Zeile_Start, Spalte_Versuchsnummer, Spalte_Versuchsname, Spalte_Ort, Spalte_Verantwortlicher, Anzahl_Zeilen] = get_Position_Versuchsliste(ws)
        # Werte einlesen wenn Zeilen gefunden wurden
        if Anzahl_Zeilen:
            for i in range(0,Anzahl_Zeilen):
                # Versuchsnummer#       Versuchsname  #      Ort   #     Verantwortlicher
                Versuchsnummer = ws[Zell_Addresse(Zeile_Start+i,Spalte_Versuchsnummer)].value
                Versuchsname = ws[Zell_Addresse(Zeile_Start+i,Spalte_Versuchsname)].value
                Ort = ws[Zell_Addresse(Zeile_Start+i,Spalte_Ort)].value
                Verantwortlicher = ws[Zell_Addresse(Zeile_Start+i,Spalte_Verantwortlicher)].value
                Versuche.append([name,Versuchsnummer, Versuchsname, Ort, Verantwortlicher ])
    return Versuche



def get_Anzahl_Gruppen_wb(wb, sheet_namen, zeile):
    Gruppen = list()
    i = 0;
    for name in sheet_namen:
        ## erste Tabelle auswählen
        ws = wb[name]
        Gruppen_Anzahl = get_Anzahl_Gruppen(ws,zeile_anfang_termine[i])
        Gruppen.append(Gruppen_Anzahl)
        i = i + 1
    return Gruppen






def get_Anzahl_Gruppen(ws, zeile):
    # eine Zeile eine Spalte über der Termintabelle sollten die Gruppennummer stehen
    # http://code.activestate.com/recipes/65117-converting-between-ascii-numbers-and-characters/
    # suche eine Zeile oberhalb in 2. Spalte nach '1', sonst 2 Zeilen oberhalb
    zeile_suchen = zeile - 2
    zeile_min = zeile_suchen - 5
    spalte_suchen = 2
    spalte_max = spalte_suchen + 8
    Gruppen = 0
    while True:
        if spalte_suchen > spalte_max:
            zeile_suchen = zeile_suchen -1
            spalte_suchen = 2
            continue
        if zeile_suchen < zeile_min:
            break
        inhalt = ws[chr(ord('A')+spalte_suchen-1)+str(zeile_suchen)].value
        if type(inhalt) != int:
            spalte_suchen = spalte_suchen + 1
            continue
        if int(inhalt) == (Gruppen + 1):
            Gruppen = Gruppen + 1
            spalte_suchen = spalte_suchen + 1
            continue
        else:
            if Gruppen > 0:
                break
            else:
                zeile_suchen = zeile_suchen -1
                continue
    return Gruppen

def get_Matrikelnamen(wb):
    ## Tabellennamen lesen
    sheet_names = wb.get_sheet_names()

    ## Tabellennamen ohne Matrikel entfernen
    sheet_namen = list()
    for sheet_name in sheet_names:
        #print(sheet_name)
        if sheet_name.find('Termine')  < 0 and sheet_name.find('Arbeitsschutz') < 0 and sheet_name.find('Tabelle1') < 0:
            #print('nicht gefunden')
            sheet_namen.append(sheet_name)

    return sheet_namen

def get_zeile_anfang_termine(wb,sheet_namen):
    ## Hier werden die Zeilennummer gespeichert wo Termine beginnen
    zeile_anfang_termine = list()
    for name in sheet_namen:
        ## erste Tabelle auswählen
        ws = wb[name]

        # Zeile mit Terminen suchen
        ## indem nach Arbeitsschutz gesucht wird oder in erster Spalte nach
        ## Versuchstermine/Pk-Gruppen, Date/ Group
        ## in Zeile n 1 - 100 suchen
        found = 0
        for i in range(4,100):
            # erste Spalte
            pruef_text = ws['A'+str(i)].value
            # wenn Zelle leer ist mit der nächsten weiter machen
            if pruef_text is None:
                continue
            if type(pruef_text) != str:
                continue
            if (len(pruef_text)<1):
                continue
    #        if (pruef_text.find("Versuchstermine")>0 or \
    #            pruef_text.find("Date/")>0 or \
            if( pruef_text.find("s. Arbeitsblatt AS-Belehrung")>= 0 or \
                pruef_text.find("Arbeitsschutzbelehrung")>= 0 or \
                pruef_text.find("Datum, Uhrzeit")>= 0):
                found = 1
                # jetzt prüfen,
                # - ob in nächster Zeile etwas steht --> weiter
                # - nächste Zeile oder
                # - beide nächste Zeile leer sind und dannach etwas steht 
                zeile = i+1 # nächste Zeile enthält Datum (oder nochmal Text)
                inhalt = ws['A'+str(zeile)].value
                if inhalt is None: # wenn die Zeile leer ist
                    # Inhalt verwerfen und nächste Zeile einlesen
                    zeile = i+2 # nächste Zeile enthält Datum (oder nochmal Text)
                    inhalt = ws['A'+str(zeile)].value
                if inhalt is None: # wenn auch die Zeile leer ist
                    # Inhalt verwerfen und nächste Zeile einlesen
                    zeile = i+3 # nächste Zeile enthält Datum (oder nochmal Text)
                    inhalt = ws['A'+str(zeile)].value
                if inhalt is None:
                    found = 0
                    continue
##                    
##                    if inhalt2 is None: # wenn die Zeile leer ist
##                if inhalt is None: # wenn die Zeile leer ist
##                    found = 0 # ist es nichts,
##                    #aber wenn in nächster Zeile doch ein Datum steht, war es nur eine Leerzeile vor der Tabelle
##                    zeile = i+1 # nächste Zeile enthält Datum (oder nochmal Text)
##                    inhalt = ws['A'+str(zeile)].value
##                    if inhalt2 is not None:
##                        found = 1
##                        break # 
##                    continue
                else:
                    if (inhalt.find("Fachbereiches")>0): # wenn dannach kein Termin
                        # sondern ein Text "des Fachbereiches Elektrotechnik für das
                        # laufende Semester bestätigt" kommt
                        found = 0
                        break # aufhören zu suchen
                    else:
                        # print(name+ " "+ str(zeile)+ " " + str(inhalt)) # Matr. Zeile und Termine
                        # Matr., Gruppen, 1. Termin
                        print(name+ ", "+ str(get_Anzahl_Gruppen(ws,zeile))+ " Gruppen, " + str(inhalt))

                break # es wurde was gefunden und es kann aufgehört werden zu suchen
            #else:
                # print("nichts gefunden")
        if found < 1:
            print("Warnung: In Tabellenblatt " + name + " wurde nichts gefunden!?")
        # abspeichern
        zeile_anfang_termine.append(zeile)
    return zeile_anfang_termine


## Dateiname festlegen und öffnen
# filename = 'PLANWS19.xlsx'
# filename = 'test1.xlsx'

filename = 'PLAN.xlsx'

wb = load_workbook(filename, read_only = True)

## Matrikelnamen aus Tabellenblättern lesen
sheet_namen = get_Matrikelnamen(wb)
    ## hier könnten Matrikel = sheet_namen entfernt
print(sheet_namen)


## Anfänge für Tabellen mit Versuchterminen finden
zeile_anfang_termine = get_zeile_anfang_termine(wb,sheet_namen)
print(zeile_anfang_termine)


Gruppen = get_Anzahl_Gruppen_wb(wb, sheet_namen, zeile_anfang_termine)
print('######## Gruppen #########')
print(Gruppen)

Termine = get_Termine(wb, sheet_namen,Gruppen,zeile_anfang_termine)
print('######## Termine #########')
#print(Termine)

Versuche = get_Versuche(wb, sheet_namen)
print('######## Versuche #########')
#print(Versuche)

Termintabelle = combine_Versuche_Termine(Versuche, Termine)
print('######## Termintabelle #########')
print(Termintabelle)

file_out_csv = 'out.csv'
export_termintabelle(Termintabelle,file_out_csv)

file_out = 'out_alle.ics'
export_ical(Termintabelle,file_out)

# xls Tabellenblatt pro Laboring. erzeugen
# ical pro Laboring. erzeugen
Laborings = get_laborings(Termintabelle)
for ing in Laborings:
    file_out = 'out_DI_' + str(ing).replace('/','') + '.ics'
    Termintabelle_gefiltert = filter_termintabelle(Termintabelle,2,ing)
    export_ical(Termintabelle_gefiltert,file_out)
    
# xls pro Prof erzeugen
# ical pro Prof. erzeugen    
Profs = get_profs(Termintabelle)
for prof in Profs:
    file_out = 'out_Prof_' + str(prof).replace('/','') + '.ics'
    Termintabelle_gefiltert = filter_termintabelle(Termintabelle,4,prof)
    export_ical(Termintabelle_gefiltert,file_out)

# xls Tabellenblatt Matrikel/Sheet_name erzeugen
# ical pro Matrikel/Sheet_name erzeugen
for matrikel in sheet_namen:
    file_out = 'out_' + str(matrikel) + '.ics'
    Termintabelle_gefiltert = filter_termintabelle(Termintabelle,1,matrikel)
    export_ical(Termintabelle_gefiltert,file_out)


# xls Tabellenblatt pro Gruppe und Matrikel erzeugen
# ical pro Gruppe und Matrikel/Sheet_name erzeugen
# xls Tabellenblatt Matrikel/Sheet_name erzeugen
# ical pro Matrikel/Sheet_name erzeugen
i = 0
for matrikel in sheet_namen:
    Anz_Gruppen = Gruppen[i]
    for gruppe in range(1,Anz_Gruppen+1):
        file_out = 'out_' + str(matrikel) + "_Gr"+ str(gruppe) + '.ics'
        Termintabelle_gefiltert1 = filter_termintabelle(Termintabelle,1,matrikel)
        Termintabelle_gefiltert2 = filter_termintabelle(Termintabelle_gefiltert1,5,gruppe)
        export_ical(Termintabelle_gefiltert2,file_out)
    i = i + 1

# ical pro Raum erzeugen?
Raeume = get_Raeume(Termintabelle)
for raum in Raeume:
    file_out = 'out_Raum_' + str(raum).replace('/','').replace('.','-') + '.ics'
    Termintabelle_gefiltert = filter_termintabelle(Termintabelle,3,raum)
    export_ical(Termintabelle_gefiltert,file_out)
